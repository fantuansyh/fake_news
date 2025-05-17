# coding: UTF-8

import os
import time
import torch
import numpy as np
import openpyxl
from openpyxl import Workbook
from need.utils import get_time_dif
from need.predicate_utils import (
    load_dataset_one, build_iterator_one,
    load_dataset_nums, build_iterator_nums,
    load_dataset_excel, load_dataset_csv
)

PAD, CLS = '[PAD]', '[CLS]'  # BERT special tokens


def _predicate(config, model, predict_iter):
    """内部预测逻辑"""
    # 加载模型权重（CPU 也能跑）
    model.load_state_dict(torch.load(config.save_path, map_location=torch.device('cpu')))
    model.eval()

    predict_all = []
    with torch.no_grad():
        for batch in predict_iter:
            outputs = model(batch)
            preds = torch.argmax(outputs, dim=1).cpu().numpy().tolist()
            predict_all.extend(preds)

    # 把 label id 转为 label 名称
    return [config.class_list[idx] for idx in predict_all]


def save_histry(new, result, save_dir='Histry/histry.xlsx'):
    """
    保存历史记录（如果目录或文件不存在，自动创建）
    new:     List[str]  — 预测的文本列表
    result:  List[str]  — 预测的分类列表
    save_dir: str       — 历史记录文件路径
    """
    # 1. 确保父目录存在
    os.makedirs(os.path.dirname(save_dir), exist_ok=True)

    def _write(ws_map, text, label):
        """向已有的 sheet 中追加一行"""
        ws = ws_map.get(label)
        if ws is None:
            ws = wb.create_sheet(title=label)
            ws.append(['content', 'channelName'])
            ws_map[label] = ws
        ws.append([text, label])

    # 2. 加载或新建 Workbook
    try:
        wb = openpyxl.load_workbook(save_dir)
    except FileNotFoundError:
        wb = Workbook()
        # 删除 openpyxl 默认创建的第一个 sheet（如果没有数据）
        if 'Sheet' in wb.sheetnames and len(wb.sheetnames) == 1:
            wb.remove(wb['Sheet'])

    # 3. 写入数据
    ws_map = {name: wb[name] for name in wb.sheetnames}
    for text, label in zip(new, result):
        _write(ws_map, text, label)

    # 4. 保存
    wb.save(save_dir)


def predicate_one(config, model, text):
    """单条文本预测"""
    print("===== Single Prediction =====")
    start = time.time()

    # 1. 构造数据迭代器
    data, new = load_dataset_one(config, text, config.pad_size)
    it = build_iterator_one(data, config)

    # 2. 预测
    labels = _predicate(config, model, it)

    print(f"Prediction: {labels[0]}")
    print("Time usage:", get_time_dif(start))

    # 3. 存历史
    save_histry([text], labels)
    return labels[0]


def predicate_nums(config, model, text_dir):
    """批量文本预测（.txt 文件）"""
    print("===== Batch Prediction (.txt) =====")
    start = time.time()

    # 1. 加载所有 .txt
    data, news = load_dataset_nums(config, text_dir, config.pad_size)
    if not data:
        print(f"目录 {text_dir} 下未找到 .txt，跳过。")
        return []

    it = build_iterator_nums(data, config)
    labels = _predicate(config, model, it)

    print("Results count:", len(labels))
    print("Time usage:", get_time_dif(start))

    # 2. 存历史 & 原目录 result.xlsx
    save_histry(news, labels)
    save_histry(news, labels, os.path.join(text_dir, 'result.xlsx'))
    return labels


def predicate_excel(config, model, excel_path):
    """Excel 文件预测"""
    print("===== Batch Prediction (Excel) =====")
    start = time.time()

    wb = openpyxl.load_workbook(excel_path)
    ws = wb.active
    data, news = load_dataset_excel(config, ws, config.pad_size)
    it = build_iterator_nums(data, config)
    labels = _predicate(config, model, it)

    print("Results count:", len(labels))
    print("Time usage:", get_time_dif(start))
    return labels


def predicate_csv(config, model, csv_path):
    """CSV 文件预测"""
    print("===== Batch Prediction (CSV) =====")
    start = time.time()

    data = load_dataset_csv(config, csv_path, config.pad_size)
    if not data:
        print(f"CSV {csv_path} 无内容，跳过。")
        return []

    it = build_iterator_nums(data, config)
    labels = _predicate(config, model, it)

    # 写回 CSV
    out_path = os.path.join('Histry', 'result.csv')
    os.makedirs(os.path.dirname(out_path), exist_ok=True)

    with open(csv_path, 'r', encoding='utf-8') as rf, \
         open(out_path, 'w', newline='', encoding='utf-8') as wf:
        reader = csv.reader(rf)
        writer = csv.writer(wf)
        header = next(reader)
        writer.writerow(header + ['class'])
        for row, label in zip(reader, labels):
            writer.writerow(row + [label])

    print("Time usage:", get_time_dif(start))
    return labels


if __name__ == '__main__':
    # —— 将 cwd 切到项目根，保证所有相对路径都能正确解析 —— #
    this_file   = os.path.abspath(__file__)
    need_dir    = os.path.dirname(this_file)          # …/news2/need
    project_root = os.path.dirname(need_dir)          # …/news2
    os.chdir(project_root)

    # —— 导入模型和配置 —— #
    from models.predicate_config import Config, Model

    config = Config()
    # 保证可复现
    np.random.seed(1)
    torch.manual_seed(1)
    torch.cuda.manual_seed_all(1)
    torch.backends.cudnn.deterministic = True

    # —— 初始化模型 —— #
    model = Model(config).to(config.device)

    # —— 单条预测 —— #
    # text = (
    #     '上海证券报＊ＳＴ中钨股票近日交易异常波动。'
    #     '２００８年１月７日，公司股票因重大无先例事项实施停牌，'
    #     '至２００８年２月１８日，经向控股股东湖南有色金属股份有限公司核实后，'
    #     '由于相关方案尚不成熟且无实质性进展，公司股票予以复牌。'
    # )
    # text = (
    #     '特斯拉在上海超级工厂发布最新款Model S Plaid，0-100公里加速仅需2.1秒，引发高性能电动车市场热议。'
    #     '国内某品牌首次推出燃油和电动混动SUV，官方称综合续航超过1000公里，定位中高端家用市场。'
    #     '欧洲某豪华车厂宣布将在2025年之前全面转型，所有新车型均将采用纯电动驱动。'
    # )
    text = (
        '国际交流项目启动，中外学生将在夏季学期互换体验，增进跨文化沟通能力。'
        '多所高校启动在线平台升级，AI智能助教参与课堂互动，为学生提供个性化学习路径'
    )
    predicate_one(config, model, text)

    # —— 批量 .txt 预测 —— #
    text_dir = 'THUCNews/pre_news'
    predicate_nums(config, model, text_dir)

    # —— 如果需要，你也可以做 Excel 或 CSV 预测 —— #
    # excel_path = 'THUCNews/data/test_4.xlsx'
    # predicate_excel(config, model, excel_path)
    # predicate_csv(config, model, 'your_news.csv')
