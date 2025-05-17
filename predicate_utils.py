# coding: UTF-8
import torch
import os
import csv
import openpyxl
from tqdm import tqdm
from need.data_cleaning import data_cleaning

PAD, CLS = '[PAD]', '[CLS]'  # padding符号, bert中综合信息符号

def load_dataset_one(config,text, pad_size=32):
    '''
    对输入的单条新闻文本进行分词
    '''
    contents = []
    news=[]
    news.append(text)
    token = config.tokenizer.tokenize(data_cleaning(text))
    token = [CLS] + token
    seq_len = len(token)
    mask = []
    token_ids = config.tokenizer.convert_tokens_to_ids(token)

    if pad_size:
        if len(token) < pad_size:
            mask = [1] * len(token_ids) + [0] * (pad_size - len(token))
            token_ids += ([0] * (pad_size - len(token)))
        else:
            mask = [1] * pad_size
            token_ids = token_ids[:pad_size]
            seq_len = pad_size
    contents.append((token_ids, seq_len, mask))
    return contents,news

def load_dataset_nums(config,text_dir, pad_size=32):
    '''
    对输入的文件夹内的所有.txt新闻文本进行分词
    '''
    contents =[]
    news=[]
    for(dirpath,dirnames,filenames) in os.walk(text_dir):
        for filename in filenames:
            if os.path.splitext(filename)[1]=='.txt':
                with open(text_dir + '/' +filename, 'r', encoding='utf-8') as f:
                    content =f.read()
                    news.append(content)
                    token_long = config.tokenizer.tokenize(data_cleaning(content))
                    token_long = [CLS] + token_long
                    seq_len = len(token_long)
                    mask = []
                    token_ids = config.tokenizer.convert_tokens_to_ids(token_long)

                    if pad_size:
                        if len(token_long) < pad_size:
                            mask = [1] * len(token_ids) + [0] * (pad_size - len(token_long))
                            token_ids += ([0] * (pad_size - len(token_long)))
                        else:
                            mask = [1] * pad_size
                            token_ids = token_ids[:pad_size]
                            seq_len = pad_size
                    contents.append((token_ids,  seq_len, mask))
    return contents,news

def load_dataset_excel(config,excel, pad_size=32):
    '''
    对输入的.xlsx文件内的新闻逐行取出并分词
    '''
    news = []
    contents =[]
    for i in tqdm(range(2, excel.max_row + 1)):
        content = str(excel.cell(i, 4).value.strip().replace('\n', '').replace('\r', ''))
        title = str(excel.cell(i, 3).value.strip().replace('\n', '').replace('\r', ''))
        new = data_cleaning(title + content)
        token_long = config.tokenizer.tokenize(new)
        token_long = [CLS] + token_long
        seq_len = len(token_long)
        mask = []
        token_ids = config.tokenizer.convert_tokens_to_ids(token_long)

        if pad_size:
            if len(token_long) < pad_size:
                mask = [1] * len(token_ids) + [0] * (pad_size - len(token_long))
                token_ids += ([0] * (pad_size - len(token_long)))
            else:
                mask = [1] * pad_size
                token_ids = token_ids[:pad_size]
                seq_len = pad_size
        news.append(content)
        contents.append((token_ids,  seq_len, mask))
    return contents,news


def load_dataset_csv(config,csv_dir, pad_size=32):
    '''
    对输入的.csv文件内的新闻逐行取出并分词
    '''
    contents =[]
    with open(csv_dir) as csvfile:
        reader = csv.DictReader(csvfile)
        for row in reader:
            content = str(row['content'].strip().replace('\n', '').replace('\r', ''))
            title = str(row['title'].strip().replace('\n', '').replace('\r', ''))
            new = data_cleaning(title + content)
            token_long = config.tokenizer.tokenize(new)
            token_long = [CLS] + token_long
            seq_len = len(token_long)
            mask = []
            token_ids = config.tokenizer.convert_tokens_to_ids(token_long)

            if pad_size:
                if len(token_long) < pad_size:
                    mask = [1] * len(token_ids) + [0] * (pad_size - len(token_long))
                    token_ids += ([0] * (pad_size - len(token_long)))
                else:
                    mask = [1] * pad_size
                    token_ids = token_ids[:pad_size]
                    seq_len = pad_size
            contents.append((token_ids,  seq_len, mask))
    return contents

class DatasetIterater_(object):
    def __init__(self, batches, batch_size, device):
        self.batch_size = batch_size
        self.batches = batches
        
        # 处理批次大小大于数据集大小的情况
        if len(batches) < batch_size:
            self.batch_size = len(batches)
        
        # 计算批次数量，确保不会出现除零错误
        self.n_batches = max(1, len(batches) // self.batch_size)
        
        self.residue = False  # 记录batch数量是否为整数
        if len(batches) % self.n_batches != 0:
            self.residue = True
        self.index = 0
        self.device = device

    def _to_tensor(self, datas):
        x = torch.LongTensor([_[0] for _ in datas]).to(self.device)

        # pad前的长度(超过pad_size的设为pad_size)
        seq_len = torch.LongTensor([_[1] for _ in datas]).to(self.device)
        mask = torch.LongTensor([_[2] for _ in datas]).to(self.device)
        return (x, seq_len, mask)

    def __next__(self):
        if self.residue and self.index == self.n_batches:
            batches = self.batches[self.index * self.batch_size: len(self.batches)]
            self.index += 1
            batches = self._to_tensor(batches)
            return batches

        elif self.index >= self.n_batches:
            self.index = 0
            raise StopIteration
        else:
            batches = self.batches[self.index * self.batch_size: (self.index + 1) * self.batch_size]
            self.index += 1
            batches = self._to_tensor(batches)
            return batches

    def __iter__(self):
        return self

    def __len__(self):
        if self.residue:
            return self.n_batches + 1
        else:
            return self.n_batches


def build_iterator_one(dataset, config):
    iter_short = DatasetIterater_(dataset, 1, config.device)
    return iter_short

def build_iterator_nums(dataset, config):
    # 根据数据集大小选择合适的批次大小
    if len(dataset) <= 5:
        batch_size = 1
    elif len(dataset) <= 10:
        batch_size = 5
    elif len(dataset) <= 20:
        batch_size = 10
    elif len(dataset) <= 32:
        batch_size = 20
    elif len(dataset) <= 64:
        batch_size = 32
    else:
        batch_size = 64
        
    iter_long = DatasetIterater_(dataset, batch_size, config.device)
    return iter_long
